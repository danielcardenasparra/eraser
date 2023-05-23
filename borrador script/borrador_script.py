import tkinter as tk
from tkinter import messagebox, simpledialog
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def agregar_datos(hl5, ip, interface, port, sap, vprn):
    
    sheet = wb.active
    orange_fill = PatternFill(start_color='fdbc00', end_color='fdbc00', fill_type='solid')
    green_fill = PatternFill(start_color='92d14d', end_color='92d14d', fill_type='solid')

    sheet['A1'].fill = orange_fill
    sheet['A2'].fill = green_fill



    sheet['A1'] = '/BORRAR SERVICIOS PREVIOS (SIN VENTANA)'
    sheet['A2'] = f"{hl5} // {ip}"
    sheet['A3'] = f"//configure port {port}"
    sheet['A4'] = "no description"
    config_1 = f'/configure service vprn {vprn} interface "{interface}" sap {port}:{sap} shutdown'
    config_2 = f'/configure service vprn {vprn} interface "{interface}" no sap {port}:{sap} shutdown'
    config_3 = f'/configure service vprn {vprn} interface "{interface}" shutdown'
    config_4 = f'/configure service vprn {vprn} no interface "{interface}"'
    
    for i, config in enumerate([config_1, config_2, config_3, config_4]):
        fila = ws.max_row + 1
        ws.cell(row=fila, column=1, value=config)
    
    wb.save('datos.xlsx')
    print('done')
    messagebox.showinfo("Datos agregados", "Los datos se han agregado correctamente.")

def limpiar_campos():
    interface_entry.delete(0, tk.END)
    sap_entry.delete(0, tk.END)
    vprn_entry.delete(0, tk.END)
    interface_entry.focus()

def nuevo_puerto():
    port = simpledialog.askstring("Nuevo Puerto", "Ingresa el número de puerto:")
    if port:
        ws.cell(row=ws.max_row + 1, column=1, value=f"//configure port {port}")
        ws.cell(row=ws.max_row + 1, column=1, value="no description")
        wb.save('datos.xlsx')
        messagebox.showinfo("Nuevo Puerto Agregado", f"El puerto {port} ha sido agregado a la hoja de cálculo.")
        port_entry.delete(0, tk.END)
        port_entry.insert(0, port)
def boton_agregar():
    hl5 = hl5_entry.get()
    ip = ip_entry.get()
    interface = interface_entry.get()
    port = port_entry.get()
    sap = sap_entry.get()
    vprn = vprn_entry.get()
    agregar_datos(hl5, ip,interface, port, sap, vprn)
    limpiar_campos()

wb = Workbook()
ws = wb.active
ws.title = "Datos"

columnas = []
for i, col in enumerate(columnas):
    ws.cell(row=1, column=i+1, value=col)

root = tk.Tk()
root.title("Agregar Datos a Excel")

hl5_label = tk.Label(root, text="HL5:")
hl5_label.grid(row=0, column=0)
hl5_entry = tk.Entry(root)
hl5_entry.grid(row=0, column=1)

 # create IP input field
ip_label = tk.Label(root, text="IP:")
ip_label.grid(row=1, column=0)
ip_entry = tk.Entry(root)
ip_entry.grid(row=1, column=1)

        # create port input field
port_label = tk.Label(root, text="Puerto:")
port_label.grid(row=2, column=0)
port_entry = tk.Entry(root)
port_entry.grid(row=2, column=1)

vprn_label = tk.Label(root, text="VPRN:")
vprn_label.grid(row=3, column=0, padx=5, pady=5)
vprn_entry = tk.Entry(root)
vprn_entry.grid(row=3, column=1, padx=5, pady=5)

# create INTERFACE input field
interface_label = tk.Label(root, text="Interface:")
interface_label.grid(row=5, column=0)
interface_entry = tk.Entry(root)
interface_entry.grid(row=5, column=1)

# create SAP listbox
sap_label = tk.Label(root, text="VLAN:")
sap_label.grid(row=6, column=0)
sap_entry = tk.Entry(root)
sap_entry.grid(row=6, column=1)


agregar_boton = tk.Button(root, text="Agregar", command=boton_agregar)
agregar_boton.grid(row=7, column=0, columnspan=2, padx=5, pady=5)
nuevo_puerto_boton = tk.Button(root, text="Nuevo Puerto", command=nuevo_puerto)
nuevo_puerto_boton.grid(row=9, column=0, columnspan=2, padx=5, pady=5)


root.mainloop()

    