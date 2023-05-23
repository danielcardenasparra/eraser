import tkinter as tk
from tkinter import messagebox, simpledialog
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


port_actualizado = False
columna_actual = 1

# Crear función para agregar datos a la hoja de cálculo
def agregar_datos(hl5, ip, interface, port, sap, vprn, columna_actual):
    global port_actualizado
    global fila
    sheet = wb.active
    orange_fill = PatternFill(start_color='fdbc00', end_color='fdbc00', fill_type='solid')
    green_fill = PatternFill(start_color='92d14d', end_color='92d14d', fill_type='solid')
    sheet.cell(row=11, column=columna_actual).fill = orange_fill
    sheet.cell(row=12, column=columna_actual).fill = green_fill
    if not port_actualizado:
        sheet.cell(row=13, column=columna_actual).value = f"//configure port {port}"
        port_actualizado = True
        primera_vez= True
    else :
        primera_vez = False
        sheet.cell(row=11, column=columna_actual).value = '/BORRAR SERVICIOS PREVIOS (SIN VENTANA)'
        sheet.cell(row=12, column=columna_actual).value = f"{hl5} // {ip}"
        primera_vez = False
# Buscar primera fila vacía en la columna actual
    fila = 11
    while ws.cell(row=fila, column=columna_actual).value is not None:
        fila += 1
    sheet.cell(row=14, column=columna_actual).value = "no description"
    config_1 = f'/configure service vprn {vprn} interface "{interface}" sap {port}:{sap} shutdown'
    config_2 = f'/configure service vprn {vprn} interface "{interface}" no sap {port}:{sap} shutdown'
    config_3 = f'/configure service vprn {vprn} interface "{interface}" shutdown'
    config_4 = f'/configure service vprn {vprn} no interface "{interface}"'
    
    for i, config in enumerate([config_1, config_2, config_3, config_4]):
        ws.cell(row=fila+i, column=columna_actual, value=config)
    
    sheet.column_dimensions[get_column_letter(columna_actual)].width = 45

    wb.save('datos.xlsx')
    print('done')
    messagebox.showinfo("Datos agregados", "Los datos se han agregado correctamente.")
    
    # Agregar imagen a la hoja de cálculo
    img = Image('telefonica.png')
    img.width = 300
    img.height = 100
    sheet.row_dimensions[1].height = 100
    sheet.add_image(img, 'A1')



# Crear función para limpiar campos de entrada después de agregar datos a la hoja de cálculo
def limpiar_campos():
    interface_entry.delete(0, tk.END)
    sap_entry.delete(0, tk.END)
    vprn_entry.delete(0, tk.END)
    interface_entry.focus()


# Crear función para agregar un nuevo puerto a la hoja de cálculo
def nuevo_puerto():
    port = simpledialog.askstring("Nuevo Puerto", "Ingresa el número de puerto:")
    if port:
        ws.cell(row=ws.max_row + 1, column=columna_actual, value=f"//configure port {port}")
        ws.cell(row=ws.max_row + 1, column=columna_actual, value="no description")
        wb.save('datos.xlsx')
        
        # Mostrar un mensaje de información indicando que el nuevo puerto ha sido agregado a la hoja de cálculo y limpiar el campo de entrada del puerto
        messagebox.showinfo("Nuevo Puerto Agregado", f"El puerto {port} ha sido agregado a la hoja de cálculo.")
        port_entry.delete(0, tk.END)
        port_entry.insert(0, port)


# Crear función para ejecutar al hacer clic en el botón "Agregar"
def boton_agregar():
    hl5 = hl5_entry.get()
    ip = ip_entry.get()
    interface = interface_entry.get()
    port = port_entry.get()
    sap = sap_entry.get()
    vprn = vprn_entry.get()
    
    agregar_datos(hl5, ip, interface, port, sap, vprn, columna_actual)
    limpiar_campos()


# Crear función para avanzar a la siguiente columna
def siguiente_columna():
    global port_actualizado
    port_actualizado = False
    global columna_actual
    columna_actual += 1
    global fila
    fila = 1
    sheet = wb.active
    messagebox.showinfo("Siguiente", f"Trabajando en la columna {get_column_letter(columna_actual)}")
    column_label.config(text=f"Columna actual: {get_column_letter(columna_actual)}")

    # Actualizar la variable ws para que apunte a la primera fila de la nueva columna
    if ws.cell(row=ws.max_row, column=columna_actual-1).value is None:
        fila = 1
        
    for i in range(1, ws.max_row + 1):
        if ws.cell(row=i, column=columna_actual).value is None:
            ws.cell(row=i, column=columna_actual).value = ""
            break

    # Limpiar los campos de entrada
    hl5_entry.delete(0, tk.END)
    ip_entry.delete(0, tk.END)
    port_entry.delete(0, tk.END)
    port_entry.insert(0, "Nuevo Puerto")
    vprn_entry.delete(0, tk.END)
    interface_entry.delete(0, tk.END)
    sap_entry.delete(0, tk.END)
    interface_entry.focus()

    # Cambiar el relleno de las celdas A11 y A12 en la columna actual
    orange_fill = PatternFill(start_color='fdbc00', end_color='fdbc00', fill_type='solid')
    green_fill = PatternFill(start_color='92d14d', end_color='92d14d', fill_type='solid')
    sheet.cell(row=11, column=columna_actual).fill = orange_fill
    sheet.cell(row=12, column=columna_actual).fill = green_fill

# Crear instancia de libro de trabajo de Excel y hoja de cálculo activa
wb = Workbook()
ws = wb.active
ws.title = "Datos"

# Crear lista vacía para las columnas y agregar encabezados a la hoja de cálculo
columnas = []
for i, col in enumerate(columnas):
    ws.cell(row=1, column=i+1, value=col)

# Crear ventana principal de la aplicación
root = tk.Tk()
root.title("Agregar Datos a Excel")

# Crear etiqueta para indicar la columna actual
column_label = tk.Label(root, text=f"Columna actual: {get_column_letter(columna_actual)}")
column_label.grid(row=0, column=0, columnspan=2)


# Crear etiqueta y entrada de texto para HL5
hl5_label = tk.Label(root, text="HL5:")
hl5_label.grid(row=1, column=0)
hl5_entry = tk.Entry(root)
hl5_entry.grid(row=1, column=1)

# Crear etiqueta y entrada de texto para la dirección IP
ip_label = tk.Label(root, text="IP:")
ip_label.grid(row=2, column=0)
ip_entry = tk.Entry(root)
ip_entry.grid(row=2, column=1)

# Crear etiqueta y entrada de texto para el número de puerto
port_label = tk.Label(root, text="Puerto:")
port_label.grid(row=3, column=0)
port_entry = tk.Entry(root)
port_entry.grid(row=3, column=1)

# Crear etiqueta y entrada de texto para VPRN
vprn_label = tk.Label(root, text="VPRN:")
vprn_label.grid(row=4, column=0, padx=5, pady=5)
vprn_entry = tk.Entry(root)
vprn_entry.grid(row=4, column=1, padx=5, pady=5)

# Crear etiqueta y entrada de texto para la interfaz
interface_label = tk.Label(root, text="Interface:")
interface_label.grid(row=5, column=0)
interface_entry = tk.Entry(root)
interface_entry.grid(row=5, column=1)

# Crear etiqueta y entrada de texto para SAP
sap_label = tk.Label(root, text="VLAN:")
sap_label.grid(row=6, column=0)
sap_entry = tk.Entry(root)
sap_entry.grid(row=6, column=1)

# Crear botón para agregar datos a la hoja de cálculo
agregar_boton = tk.Button(root, text="Agregar", command=boton_agregar)
agregar_boton.grid(row=7, column=0, columnspan=2, padx=5, pady=5)

# Crear botón para agregar nuevos puertos a la hoja de cálculo
nuevo_puerto_boton = tk.Button(root, text="Nuevo Puerto", command=nuevo_puerto)
nuevo_puerto_boton.grid(row=8, column=0, columnspan=2, padx=5, pady=5)

# Crear botón para avanzar a la siguiente columna
siguiente_boton = tk.Button(root, text="Siguiente", command=siguiente_columna)
siguiente_boton.grid(row=9, column=0, columnspan=2, padx=5, pady=5)

# Colocar la ventana en el bucle principal de la aplicación
root.mainloop()

